"""XLSX export for error-checked annotations with crop thumbnails."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font
from PIL import Image as PILImage


def make_thumbnail(jpeg_bytes, height=80):
    """Resize JPEG bytes to thumbnail, return BytesIO + dimensions."""
    img = PILImage.open(BytesIO(jpeg_bytes))
    ratio = height / img.height
    new_w = max(1, int(img.width * ratio))
    img = img.resize((new_w, height), PILImage.LANCZOS)
    buf = BytesIO()
    img.save(buf, format="JPEG", quality=85)
    buf.seek(0)
    return buf, new_w, height


def export_errors_xlsx(dataset, ann_ids):
    """Generate XLSX with error items and embedded crop thumbnails.

    Returns BytesIO containing the XLSX file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "errors"

    headers = ["barcode", "shipment_id", "category", "annotation_id",
               "frame", "source_image", "crop_image"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 55
    ws.column_dimensions["G"].width = 20

    THUMB_HEIGHT = 80
    ROW_HEIGHT = THUMB_HEIGHT * 0.75

    for i, ann_id in enumerate(ann_ids):
        row = i + 2
        info = dataset.get_annotation_info(ann_id)
        if info is None:
            continue

        ws.cell(row=row, column=1, value=info["barcode"])
        ws.cell(row=row, column=2, value=info["shipment_id"])
        ws.cell(row=row, column=3, value=info["category"])
        ws.cell(row=row, column=4, value=ann_id)
        ws.cell(row=row, column=5, value=info.get("frame", -1))
        ws.cell(row=row, column=6, value=info["image_file"])

        # Embed crop thumbnail
        crop_bytes = dataset.get_crop_jpeg(ann_id)
        if crop_bytes:
            try:
                buf, tw, th = make_thumbnail(crop_bytes, THUMB_HEIGHT)
                img = XlImage(buf)
                img.width = tw
                img.height = th
                ws.add_image(img, f"G{row}")
            except Exception:
                ws.cell(row=row, column=7, value="[thumbnail error]")
        else:
            ws.cell(row=row, column=7, value="[no crop]")

        ws.row_dimensions[row].height = ROW_HEIGHT

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_memos_xlsx(dataset, memos):
    """Generate XLSX with review memos and embedded crop thumbnails.

    Args:
        dataset: CocoDataset instance
        memos: dict of memo_id -> memo dict

    Returns BytesIO containing the XLSX file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "review_memos"

    headers = ["No", "frame_index", "file_name", "category", "annotation_id",
               "barcode", "shipment_id", "error_description", "timestamp", "crop_image"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 20

    THUMB_HEIGHT = 80
    ROW_HEIGHT = THUMB_HEIGHT * 0.75

    sorted_memos = sorted(memos.values(),
                          key=lambda m: (m.get("frame_index", 0), m.get("ann_id", 0)))

    for i, memo in enumerate(sorted_memos):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=memo.get("frame_index", 0))
        ws.cell(row=row, column=3, value=memo.get("image_file", ""))
        ws.cell(row=row, column=4, value=memo.get("category", ""))
        ws.cell(row=row, column=5, value=memo.get("ann_id", ""))
        ws.cell(row=row, column=6, value=memo.get("barcode", ""))
        ws.cell(row=row, column=7, value=memo.get("shipment_id", ""))
        ws.cell(row=row, column=8, value=memo.get("text", ""))
        ws.cell(row=row, column=9, value=memo.get("timestamp", ""))

        if memo.get("type") == "missing":
            xy = memo.get("click_xy", (0, 0))
            ws.cell(row=row, column=10, value=f"[missing @ ({xy[0]},{xy[1]})]")
        else:
            ann_id = memo.get("ann_id")
            if ann_id is not None:
                crop_bytes = dataset.get_crop_jpeg(int(ann_id))
                if crop_bytes:
                    try:
                        buf, tw, th = make_thumbnail(crop_bytes, THUMB_HEIGHT)
                        img = XlImage(buf)
                        img.width = tw
                        img.height = th
                        ws.add_image(img, f"J{row}")
                    except Exception:
                        ws.cell(row=row, column=10, value="[thumbnail error]")
                else:
                    ws.cell(row=row, column=10, value="[no crop]")

        ws.row_dimensions[row].height = ROW_HEIGHT

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
